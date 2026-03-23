from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
from datetime import datetime, date, timedelta
from io import BytesIO
from sqlalchemy import func
from models import db, Booking, User, Hall, Notification

admin_bp = Blueprint('admin', __name__)

def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return wrapped

def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('auth.login'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('booking.dashboard'))
        return f(*args, **kwargs)
    return wrapped


@admin_bp.route('/pending')
@login_required
@admin_required
def pending_approvals():
    pending = Booking.query.filter_by(status='pending').order_by(Booking.created_at.asc()).all()
    return render_template('admin_pending.html', bookings=pending)


@admin_bp.route('/approve/<int:booking_id>', methods=['POST'])
@login_required
@admin_required
def approve_booking(booking_id):
    b = Booking.query.get_or_404(booking_id)
    if b.status != 'pending':
        flash('Booking is not pending approval.', 'error')
        return redirect(url_for('admin.pending_approvals'))
    notes = (request.form.get('admin_notes') or '').strip()
    b.status = 'approved'
    b.admin_notes = notes
    # Notify the booking owner
    db.session.add(Notification(
        user_id=b.user_id,
        message=f'Your booking for "{b.hall.name if b.hall else "Hall"}" on {b.booking_date} was approved.'
    ))
    db.session.commit()
    flash('Booking approved.', 'success')
    return redirect(url_for('admin.pending_approvals'))


@admin_bp.route('/reject/<int:booking_id>', methods=['POST'])
@login_required
@admin_required
def reject_booking(booking_id):
    b = Booking.query.get_or_404(booking_id)
    if b.status != 'pending':
        flash('Booking is not pending approval.', 'error')
        return redirect(url_for('admin.pending_approvals'))
    reason = (request.form.get('admin_notes') or request.form.get('reject_reason') or '').strip()
    if not reason:
        flash('Rejection reason is required.', 'error')
        return redirect(url_for('admin.pending_approvals'))
    b.status = 'rejected'
    b.admin_notes = reason
    # Notify the booking owner
    db.session.add(Notification(
        user_id=b.user_id,
        message=f'Your booking for "{b.hall.name if b.hall else "Hall"}" on {b.booking_date} was rejected. Reason: {reason}'
    ))
    db.session.commit()
    flash('Booking rejected.', 'success')
    return redirect(url_for('admin.pending_approvals'))


@admin_bp.route('/reports')
@login_required
@admin_required
def reports():
    now = date.today()
    first = date(now.year, now.month, 1)
    return render_template('admin_reports.html', from_date=first.isoformat(), to_date=now.isoformat())


@admin_bp.route('/cancellations')
@login_required
@admin_required
def cancellation_history():
    """List all cancelled bookings with reason and date."""
    cancelled = Booking.query.filter_by(status='cancelled').order_by(Booking.cancelled_at.desc()).all()
    return render_template('admin_cancellations.html', bookings=cancelled)


@admin_bp.route('/staff', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_staff():
    """Allow admin to create staff accounts using username + faculty name + password."""
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        full_name = (request.form.get('full_name') or '').strip()
        password = request.form.get('password') or ''
        if not username or not full_name or not password:
            flash('Username, faculty name, and password are required.', 'error')
        else:
            existing = User.query.filter(func.lower(User.username) == username.lower()).first()
            if existing:
                flash('This username already exists. Please choose another username.', 'error')
            else:
                staff = User(username=username, full_name=full_name, role='staff')
                staff.set_password(password)
                db.session.add(staff)
                db.session.commit()
                flash('Staff created successfully.', 'success')
                return redirect(url_for('admin.manage_staff'))
    staff_users = User.query.filter_by(role='staff').order_by(User.username.asc()).all()
    return render_template('admin_staff.html', staff_users=staff_users)


@admin_bp.route('/staff/edit/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def edit_staff(user_id):
    u = User.query.get_or_404(user_id)
    if u.role != 'staff':
        flash('Only staff users can be edited here.', 'error')
        return redirect(url_for('admin.manage_staff'))

    new_username = (request.form.get('username') or '').strip()
    new_name = (request.form.get('full_name') or '').strip()
    new_password = request.form.get('password') or ''

    if not new_username or not new_name:
        flash('Username and faculty name are required.', 'error')
        return redirect(url_for('admin.manage_staff'))

    existing = User.query.filter(User.id != u.id, func.lower(User.username) == new_username.lower()).first()
    if existing:
        flash('Another staff already uses this username.', 'error')
        return redirect(url_for('admin.manage_staff'))

    u.full_name = new_name
    u.username = new_username
    if new_password.strip():
        u.set_password(new_password)
    db.session.commit()
    flash('Staff updated successfully.', 'success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/staff/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_staff(user_id):
    u = User.query.get_or_404(user_id)
    if u.role != 'staff':
        flash('Only staff users can be deleted here.', 'error')
        return redirect(url_for('admin.manage_staff'))

    has_bookings = Booking.query.filter_by(user_id=u.id).first() is not None
    if has_bookings:
        flash('Cannot delete this staff because bookings exist for this account.', 'error')
        return redirect(url_for('admin.manage_staff'))

    # Remove their notifications (if any) before deleting the user
    Notification.query.filter_by(user_id=u.id).delete()
    db.session.delete(u)
    db.session.commit()
    flash('Staff deleted successfully.', 'success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/staff/delete-all', methods=['POST'])
@login_required
@admin_required
def delete_all_staff():
    """Remove all staff users. Bookings are reassigned to admin."""
    admin_user = User.query.filter_by(role='admin').first()
    if not admin_user:
        flash('Admin user not found. Cannot delete staff.', 'error')
        return redirect(url_for('admin.manage_staff'))

    staff_users = User.query.filter_by(role='staff').all()
    count = 0
    for u in staff_users:
        # Reassign bookings to admin so history is preserved
        Booking.query.filter_by(user_id=u.id).update({'user_id': admin_user.id})
        Notification.query.filter_by(user_id=u.id).delete()
        db.session.delete(u)
        count += 1
    db.session.commit()
    flash(f'All {count} staff removed. You can now create new staff.', 'success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/classes', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_classes():
    flash('Class creation module has been removed.', 'info')
    return redirect(url_for('booking.dashboard'))


@admin_bp.route('/classes/delete/<int:class_id>', methods=['POST'])
@login_required
@admin_required
def delete_class(class_id):
    flash('Class creation module has been removed.', 'info')
    return redirect(url_for('booking.dashboard'))


@admin_bp.route('/reports/monthly')
@login_required
@admin_required
def reports_monthly():
    from_date_s = request.args.get('from_date') or ''
    to_date_s = request.args.get('to_date') or ''
    fmt = request.args.get('format', 'html')  # html, pdf, excel
    status = (request.args.get('status') or 'all').strip().lower()
    page = request.args.get('page', type=int) or 1
    per_page = 10
    if not from_date_s or not to_date_s:
        flash('From date and To date are required.', 'error')
        return redirect(url_for('admin.reports'))
    try:
        start = datetime.strptime(from_date_s, '%Y-%m-%d').date()
        end = datetime.strptime(to_date_s, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        flash('Invalid date format. Use YYYY-MM-DD.', 'error')
        return redirect(url_for('admin.reports'))
    if end < start:
        flash('To date must be on or after From date.', 'error')
        return redirect(url_for('admin.reports'))
    allowed = {'approved', 'pending', 'cancelled', 'rejected'}
    visible_statuses = list(allowed) if status == 'all' else ([status] if status in allowed else list(allowed))

    q = Booking.query.filter(
        Booking.status.in_(visible_statuses),
        Booking.booking_date >= start,
        Booking.booking_date <= end
    ).order_by(Booking.booking_date.desc(), Booking.created_at.desc())

    # Counts for date range (all statuses)
    count_pending = Booking.query.filter(Booking.booking_date >= start, Booking.booking_date <= end, Booking.status == 'pending').count()
    count_approved = Booking.query.filter(Booking.booking_date >= start, Booking.booking_date <= end, Booking.status == 'approved').count()
    count_cancelled = Booking.query.filter(Booking.booking_date >= start, Booking.booking_date <= end, Booking.status == 'cancelled').count()
    count_rejected = Booking.query.filter(Booking.booking_date >= start, Booking.booking_date <= end, Booking.status == 'rejected').count()
    count_all = count_pending + count_approved + count_cancelled + count_rejected
    counts = {'pending': count_pending, 'approved': count_approved, 'cancelled': count_cancelled, 'rejected': count_rejected, 'all': count_all}

    # Exports should include all matching rows (no pagination).
    if fmt in ('excel', 'pdf'):
        bookings = q.all()
    else:
        bookings_pagination = q.paginate(page=page, per_page=per_page, error_out=False)
        bookings = bookings_pagination.items
    if fmt == 'excel':
        return _export_excel(bookings, start, end, counts)
    if fmt == 'pdf':
        try:
            return _export_pdf(bookings, start, end, counts)
        except ModuleNotFoundError as e:
            if getattr(e, "name", "") == "reportlab":
                flash('PDF export requires "reportlab". Install it with: pip install -r requirements.txt', 'error')
                return redirect(url_for('admin.reports_monthly', from_date=from_date_s, to_date=to_date_s, status=status, format='html'))
            raise
    return render_template(
        'admin_reports_monthly.html',
        bookings=bookings,
        from_date=from_date_s,
        to_date=to_date_s,
        status=status,
        pagination=bookings_pagination,
        counts=counts,
    )


def _fmt_date_excel(d, fmt='%d-%m-%Y'):
    if d is None:
        return ''
    try:
        if hasattr(d, 'strftime'):
            return d.strftime(fmt)
    except (ValueError, TypeError):
        pass
    return str(d) if d else ''


def _export_excel(bookings, start_date, end_date, counts=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    counts = counts or {}
    wb = Workbook()
    ws = wb.active
    label = f'{start_date} to {end_date}' if start_date != end_date else str(start_date)
    ws.title = label[:31]
    ws['A1'] = f'Hall Usage Report - {label}'
    ws['A1'].font = Font(bold=True, size=14)
    headers = ['S.No', 'Event From', 'Event To', 'Hall', 'Department', 'Purpose', 'Status', 'Faculty Name', 'Booked On']
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
        ws.cell(row=3, column=c).font = Font(bold=True)
    for r, b in enumerate(bookings, 4):
        sn = r - 3
        created_local = getattr(b, "created_at_local", None) or b.created_at
        ws.cell(row=r, column=1, value=sn)
        ws.cell(row=r, column=2, value=_fmt_date_excel(b.booking_date))
        ws.cell(row=r, column=3, value=_fmt_date_excel(b.event_end_date))
        ws.cell(row=r, column=4, value=b.hall.name if b.hall else '')
        ws.cell(row=r, column=5, value=b.department)
        ws.cell(row=r, column=6, value=(b.purpose or '')[:100])
        ws.cell(row=r, column=7, value=(b.status or '').capitalize())
        ws.cell(row=r, column=8, value=(b.user.full_name if b.user and b.user.full_name else (b.user.username if b.user else '')))
        ws.cell(row=r, column=9, value=_fmt_date_excel(created_local))
    last_row = 4 + len(bookings) - 1 if bookings else 3
    if counts:
        ws.cell(row=last_row + 2, column=1, value=f"Summary: All: {counts.get('all', 0)} | Pending: {counts.get('pending', 0)} | Approved: {counts.get('approved', 0)} | Rejected: {counts.get('rejected', 0)} | Cancelled: {counts.get('cancelled', 0)}")
        ws.cell(row=last_row + 2, column=1).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'hall_usage_{start_date}_to_{end_date}.xlsx'.replace(' ', '_')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fn, as_attachment=True)


def _fmt_date(d, fmt='%d-%m-%Y'):
    """Format date/datetime for PDF (DD-MM-YYYY); returns '-' if None."""
    if d is None:
        return '-'
    try:
        if hasattr(d, 'strftime'):
            return d.strftime(fmt)
    except (ValueError, TypeError):
        pass
    return str(d) if d else '-'


def _export_pdf(bookings, start_date, end_date, counts=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from xml.sax.saxutils import escape
    counts = counts or {}
    buf = BytesIO()
    start_str = _fmt_date(start_date)
    end_str = _fmt_date(end_date)
    label = f'{start_str} to {end_str}' if start_date != end_date else start_str
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12, rightMargin=12, topMargin=14, bottomMargin=14)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        name="cell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        spaceAfter=0,
        spaceBefore=0,
        wordWrap="CJK",
        splitLongWords=1,
    )
    dept_style = ParagraphStyle(
        name="dept",
        parent=cell_style,
        fontSize=6.5,
        leading=8,
        wordWrap="CJK",
        splitLongWords=1,
    )
    header_style = ParagraphStyle(
        name="header",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        spaceAfter=0,
        spaceBefore=0,
    )
    story = [Paragraph(f'Hall Usage Report - {label}', styles['Title']), Spacer(1, 0.3 * inch)]
    data = [[
        Paragraph('S.No', header_style),
        Paragraph('Event From', header_style),
        Paragraph('Event To', header_style),
        Paragraph('Hall', header_style),
        Paragraph('Department', header_style),
        Paragraph('Purpose', header_style),
        Paragraph('Status', header_style),
        Paragraph('Faculty Name', header_style),
        Paragraph('Booked On', header_style),
    ]]
    for idx, b in enumerate(bookings, 1):
        purpose = Paragraph(escape(b.purpose or ''), cell_style)
        dept = Paragraph(escape(b.department or ''), dept_style)
        hall_name = Paragraph(escape(b.hall.name if b.hall else ''), cell_style)
        event_from_str = _fmt_date(b.booking_date) if b.booking_date else '-'
        event_to_str = _fmt_date(b.event_end_date) if b.event_end_date else '-'
        created_local = getattr(b, "created_at_local", None)
        booked_on_str = _fmt_date(created_local) if created_local else '-'
        faculty = b.user.full_name if (b.user and b.user.full_name) else (b.user.username if b.user else '-')
        data.append([
            str(idx),
            event_from_str,
            event_to_str,
            hall_name,
            dept,
            purpose,
            (b.status or '').capitalize(),
            faculty,
            booked_on_str,
        ])
    t = Table(
        data,
        repeatRows=1,
        colWidths=[
            0.50*inch,  # S.No
            1.00*inch,  # Event From
            1.00*inch,  # Event To
            1.15*inch,  # Hall
            1.80*inch,  # Department
            1.90*inch,  # Purpose
            0.75*inch,  # Status
            1.10*inch,  # Faculty Name
            1.00*inch,  # Booked On
        ],
    )
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('WORDWRAP', (0, 0), (-1, -1), True),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    if counts:
        summary = f"Summary: All: {counts.get('all', 0)} | Pending: {counts.get('pending', 0)} | Approved: {counts.get('approved', 0)} | Rejected: {counts.get('rejected', 0)} | Cancelled: {counts.get('cancelled', 0)}"
        story.append(Spacer(1, 0.2 * inch))
        story.append(Paragraph(summary, ParagraphStyle(name="summary", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold")))
    doc.build(story)
    buf.seek(0)
    fn = f'hall_usage_{start_date}_to_{end_date}.pdf'.replace(' ', '_')
    return send_file(buf, mimetype='application/pdf',
                     download_name=fn, as_attachment=True)


